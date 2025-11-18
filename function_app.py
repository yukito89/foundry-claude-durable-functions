import azure.functions as func
import logging
import io
import pandas as pd
from openpyxl import load_workbook
import re
import zipfile
from openai import AzureOpenAI
from urllib.parse import quote
from pathlib import Path
import os
from dotenv import load_dotenv
import boto3
from botocore.config import Config
import json
import time
from prompts import (
    STRUCTURING_PROMPT, 
    EXTRACT_TEST_PERSPECTIVES_PROMPT, 
    CREATE_TEST_SPEC_PROMPT_SIMPLE,
    CREATE_TEST_SPEC_PROMPT_DETAILED,
    DIFF_DETECTION_PROMPT,
    EXTRACT_TEST_PERSPECTIVES_PROMPT_WITH_DIFF,
    CREATE_TEST_SPEC_PROMPT_WITH_DIFF
)

# .envファイルから環境変数を読み込む
load_dotenv()

# FunctionAppの初期化（IP制限を使用するため認証レベルをANONYMOUSに変更）
app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

# --- LLMサービス設定 ---
llm_service = os.getenv("LLM_SERVICE", "AWS")

# --- Azure OpenAI Service 接続情報 ---
azure_api_key = os.getenv("AZURE_OPENAI_API_KEY")
azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")
azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")

# --- AWS Bedrock 接続情報 ---
aws_region = os.getenv("AWS_REGION")
aws_access_key_id = os.getenv("AWS_ACCESS_KEY_ID")
aws_secret_access_key = os.getenv("AWS_SECRET_ACCESS_KEY")
aws_bedrock_model_id = os.getenv("AWS_BEDROCK_MODEL_ID")

# LLMクライアントの初期化用変数
azure_client = None
bedrock_client = None

# 必須環境変数のチェック関数
def validate_env():
    if llm_service == "AZURE":
        # Azure OpenAIに必要な環境変数がすべて設定されているか確認
        required = [azure_api_key, azure_endpoint, azure_api_version, azure_deployment]
        if not all(required):
            raise ValueError("Azure OpenAI の必須環境変数が設定されていません。")
    elif llm_service == "AWS":
        # AWS Bedrockに必要な環境変数がすべて設定されているか確認
        required = [aws_region, aws_access_key_id, aws_secret_access_key, aws_bedrock_model_id]
        if not all(required):
            raise ValueError("AWS Bedrock の必須環境変数が設定されていません。")
    else:
        # サポートされていないLLMサービスが指定された場合のエラー
        raise ValueError(f"無効なLLMサービスが指定されました: {llm_service}")

# LLMクライアントの初期化関数
def initialize_client():
    global azure_client, bedrock_client
    validate_env()  # 環境変数の妥当性をチェック

    if llm_service == "AZURE":
        # Azure OpenAIクライアントの初期化
        azure_client = AzureOpenAI(
            api_version=azure_api_version,
            azure_endpoint=azure_endpoint,
            api_key=azure_api_key,
        )
    elif llm_service == "AWS":
        # AWS Bedrockクライアントの初期化（タイムアウト設定付き）
        config = Config(read_timeout=600, connect_timeout=60)
        bedrock_client = boto3.client(
            "bedrock-runtime",
            region_name=aws_region,
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key,
            config=config,
        )

# LLMサービスを呼び出す共通関数
def call_llm(system_prompt: str, user_prompt: str, max_retries: int = 10) -> str:
    """
    指定されたLLMサービス（AzureまたはAWS）を使ってプロンプトを送信し、応答を取得する。
    system_prompt: システムプロンプト（モデルの振る舞いを定義）
    user_prompt: ユーザーからの入力
    max_retries: 最大リトライ回数
    戻り値: モデルからの応答テキスト
    """
    global azure_client, bedrock_client
    
    # クライアントが未初期化の場合は初期化する
    if llm_service == "AZURE" and azure_client is None:
        initialize_client()
    elif llm_service == "AWS" and bedrock_client is None:
        initialize_client()
    
    for attempt in range(max_retries):
        try:
            if llm_service == "AZURE":
                # Azure OpenAIにチャット形式でリクエストを送信
                response = azure_client.chat.completions.create(
                    model=azure_deployment,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    max_completion_tokens=32768,
                )
                return response.choices[0].message.content

            elif llm_service == "AWS":
                # AWS BedrockにConverse APIでリクエストを送信
                response = bedrock_client.converse(
                    modelId=aws_bedrock_model_id,
                    messages=[{"role": "user", "content": [{"text": user_prompt}]}],
                    system=[{"text": system_prompt}],
                    inferenceConfig={"maxTokens": 64000},
                )
                # レスポンスの構造を確認してから取得
                if 'output' in response and 'message' in response['output']:
                    return response['output']['message']['content'][0]['text']
                else:
                    logging.error(f"予期しないレスポンス構造: {json.dumps(response, ensure_ascii=False)}")
                    raise RuntimeError("AWS Bedrockからの応答形式が不正です。")

        except Exception as e:
            error_message = str(e)
            # ThrottlingExceptionの場合はリトライ
            if "ThrottlingException" in error_message or "Too many requests" in error_message:
                if attempt < max_retries - 1:
                    # エクスポネンシャルバックオフ
                    wait_time = min((2 ** attempt) * 3 + (attempt * 5), 120)  # 最大120秒
                    logging.warning(f"{llm_service} API レート制限エラー。{wait_time}秒後にリトライします（{attempt + 1}/{max_retries}）")
                    time.sleep(wait_time)
                    continue
                else:
                    logging.error(f"{llm_service} API呼び出しが最大リトライ回数に達しました")
                    raise RuntimeError(f"{llm_service} APIのレート制限エラー。しばらく待ってから再試行してください。")
            else:
                # その他のエラーは即座に失敗
                logging.error(f"{llm_service} API呼び出し中にエラーが発生しました: {error_message}")
                raise RuntimeError(f"{llm_service} API呼び出しに失敗しました: {error_message}")
    
    raise RuntimeError(f"{llm_service} API呼び出しに失敗しました")


def structuring(prompt: str) -> str:
    return call_llm(STRUCTURING_PROMPT, prompt)

def extract_test_perspectives(prompt: str) -> str:
    return call_llm(EXTRACT_TEST_PERSPECTIVES_PROMPT, prompt)

def create_test_spec(prompt: str, granularity: str = "simple") -> str:
    """テスト仕様書を生成
    
    Args:
        prompt: 設計書とテスト観点を含むプロンプト
        granularity: 粒度 ("simple" or "detailed")
    """
    if granularity == "detailed":
        return call_llm(CREATE_TEST_SPEC_PROMPT_DETAILED, prompt)
    else:
        return call_llm(CREATE_TEST_SPEC_PROMPT_SIMPLE, prompt)

def process_excel_to_markdown(files) -> str:
    """Excelファイル群をMarkdownに変換
    
    Args:
        files: アップロードされたExcelファイルのリスト
    Returns:
        統合されたMarkdown文字列
    """
    all_md_sheets = []
    all_toc_list = []
    
    for file in files:
        file_bytes = file.read()
        filename = file.filename
        excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
        
        for sheet_name, df in excel_data.items():
            full_sheet_name = f"{Path(filename).stem}_{sheet_name}"
            anchor = re.sub(r'[^a-z0-9-]', '', full_sheet_name.strip().lower().replace(' ', '-'))
            all_toc_list.append(f'- [{full_sheet_name}](#{anchor})')
            
            sheet_content = f"## {full_sheet_name}\n\n"
            raw_text = '\n'.join(df.apply(lambda row: ' | '.join(row.astype(str).fillna('')), axis=1))
            structuring_prompt = f'--- Excelシート「{full_sheet_name}」 ---\n{raw_text}'
            structured_content = structuring(structuring_prompt)
            sheet_content += structured_content
            all_md_sheets.append(sheet_content)
    
    md_output = "# 詳細設計書\n\n## 目次\n\n"
    md_output += "\n".join(all_toc_list)
    md_output += "\n\n---\n\n"
    md_output += "\n\n---\n\n".join(all_md_sheets)
    return md_output



@app.route(route="upload", methods=["POST"])
def upload(req: func.HttpRequest) -> func.HttpResponse:
    try:
        # 複数のExcelファイルを受け取る（getlistで複数ファイルに対応）
        files = req.files.getlist("documentFiles")
        if not files:
            return func.HttpResponse("ファイルがアップロードされていません", status_code=400)
        
        # ファイル数制限（最大10ファイル）
        MAX_FILES = 10
        if len(files) > MAX_FILES:
            return func.HttpResponse(f"ファイル数は{MAX_FILES}件までです", status_code=400)
        
        # ファイルサイズとタイプのチェック
        MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
        for file in files:
            if not file.filename.endswith('.xlsx'):
                return func.HttpResponse("Excelファイル(.xlsx)のみ対応しています", status_code=400)
            
            # ファイルサイズチェック
            file.stream.seek(0, 2)  # ファイル末尾に移動
            file_size = file.stream.tell()
            file.stream.seek(0)  # 先頭に戻す
            if file_size > MAX_FILE_SIZE:
                return func.HttpResponse("ファイルサイズが大きすぎます（最大50MB）", status_code=400)
        
        # 粒度パラメータを取得（デフォルトは"simple"）
        granularity = req.form.get("granularity", "simple")
        if granularity not in ["simple", "detailed"]:
            granularity = "simple"
        logging.info(f"テスト仕様書の粒度: {granularity}")
            
    except Exception as e:
        logging.error(f"ファイル取得エラー: {e}")
        return func.HttpResponse("ファイルの取得に失敗しました", status_code=400)

    logging.info(f"{len(files)}件のファイルを受信しました。単体テスト生成を開始します。")

    try:
        # 複数ファイルの全シートを統合するためのリストを初期化
        all_md_sheets = []  # 全ファイルの全シートのMarkdownコンテンツを格納
        all_toc_list = []   # 全ファイルの全シートの目次エントリを格納
        
        # アップロードされた各Excelファイルを順に処理
        for file in files:
            file_bytes = file.read()  # ファイルをバイナリデータとして読み込み
            filename = file.filename   # ファイル名を取得
            logging.info(f"{filename} を処理中...")
            
            # Excelファイルをメモリ上で読み込み、全シートを辞書形式で取得
            # sheet_name=None: 全シートを読み込む
            # header=None: ヘッダー行を自動認識せず、全行をデータとして扱う
            excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)

            # 各ファイルの各シートを処理
            for sheet_name, df in excel_data.items():
                # ファイル名とシート名を組み合わせて一意な名前を生成
                # 例: "設計書1.xlsx"の"画面仕様"シート → "設計書1_画面仕様"
                full_sheet_name = f"{Path(filename).stem}_{sheet_name}"
                
                # Markdownのアンカーリンク用のIDを生成（GitHub-flavored Markdown形式）
                # 英数字とハイフン以外を削除し、小文字化
                anchor = re.sub(r'[^a-z0-9-]', '', full_sheet_name.strip().lower().replace(' ', '-'))
                
                # 目次にエントリを追加（Markdownリンク形式）
                all_toc_list.append(f'- [{full_sheet_name}](#{anchor})')
                
                # シートのコンテンツをMarkdown見出しで開始
                sheet_content = f"## {full_sheet_name}\n\n"

                # AIを使ってシートの生データを構造化されたMarkdownに変換
                logging.info(f"「{full_sheet_name}」シートをAIで構造化します。")
                try:
                    # DataFrameの各行をパイプ区切りのテキストに変換
                    # 例: "A1 | B1 | C1\nA2 | B2 | C2"
                    raw_text = '\n'.join(df.apply(lambda row: ' | '.join(row.astype(str).fillna('')), axis=1))
                    
                    # AIに送信するプロンプトを作成
                    structuring_prompt = f'''
                        --- Excelシート「{full_sheet_name}」 ---
                        {raw_text}
                    '''
                    # AIに構造化を依頼（STRUCTURING_PROMPTはシステムプロンプト）
                    structured_content = structuring(structuring_prompt)
                    sheet_content += structured_content
                    
                except Exception as e:
                    # AI処理が失敗した場合はエラーメッセージを追加
                    logging.error(f"AIによるシート構造化中にエラー: {e}")
                    sheet_content += "（AIによる構造化に失敗しました）"
                
                # 構造化されたシートコンテンツをリストに追加
                all_md_sheets.append(sheet_content)

        # --- 1. 全ファイルの全シートを結合して最終的なMarkdown設計書を生成 ---
        logging.info("全ファイルの処理が完了。最終的な設計書を組み立てます。")
        # タイトルを追加（複数ファイルの場合は汎用的な名前を使用）
        md_output_first = f"# 詳細設計書\n\n"
        # 目次セクションを追加
        md_output_first += "## 目次\n\n"
        # 全シートの目次エントリを結合
        md_output_first += "\n".join(all_toc_list)
        md_output_first += "\n\n---\n\n"
        # 全シートのコンテンツを区切り線で区切って結合
        md_output_first += "\n\n---\n\n".join(all_md_sheets)
        logging.info("Markdown設計書をメモリ上に生成しました。")
        
        # --- 2. AIによるテスト観点抽出 ---
        logging.info("設計書全体をAIに渡し、テスト観点を抽出します。")
        extract_test_perspectives_prompt = f'''
            --- 設計書 ---
            {md_output_first}
        '''
        md_output_second = extract_test_perspectives(extract_test_perspectives_prompt)
        logging.info("テスト観点抽出が完了し、メモリ上に保持しました。")

        # --- 3. AIによるテスト仕様書生成 ---
        logging.info("設計書全体をAIに渡し、テスト仕様書を生成します。")
        test_gen_prompt = f'''
            --- 設計書 ---
            {md_output_first}
            
            --- テスト観点 ---
            {md_output_second}
        '''

        md_output_third = create_test_spec(test_gen_prompt, granularity)
        logging.info("テスト仕様書の生成が完了し、メモリ上に保持しました。")

        # --- 4. テスト仕様書(Markdown)をExcelに変換 ---
        logging.info("テスト仕様書をMarkdownからExcel形式に変換します。")
        
        # Markdownから表部分だけを抽出
        md_lines = [line.strip() for line in md_output_third.splitlines() if line.strip().startswith("|")]
        
        if not md_lines:
            logging.error("テスト仕様書にMarkdown表が見つかりませんでした")
            return func.HttpResponse("テスト仕様書の生成に失敗しました（表形式が見つかりません）", status_code=500)
        
        # Markdown表をTSV形式に変換
        # 例: "| No | 大区分 | 中区分 |" → "No\t大区分\t中区分"
        # 1. strip("|")で行頭・行末の"|"を削除
        # 2. replace("|", "\t")で列区切りの"|"をタブに変換
        tsv_text = "\n".join([line.strip("|").replace("|", "\t") for line in md_lines])

        # TSVテキストをDataFrameに変換
        df = pd.read_csv(io.StringIO(tsv_text), sep="\t")
        # 列名の前後の空白を削除
        df.columns = [col.strip() for col in df.columns]
        
        # 必須列の存在確認（6列構成）
        required_columns = ["No", "大区分", "中区分", "テストケース", "期待結果", "参照元"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logging.error(f"必須列が不足しています: {missing_columns}")
            return func.HttpResponse(f"テスト仕様書の形式が不正です（不足列: {', '.join(missing_columns)}）", status_code=500)
        
        # 既存テンプレートを読み込み
        template_path = "単体テスト仕様書.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        # マッピング定義（DataFrame列名 → Excel列番号）
        column_map = {
            "No": 1,           # A列
            "大区分": 2,       # B列
            "中区分": 6,       # F列
            "テストケース": 10, # J列
            "期待結果": 23,     # W列
            "参照元": 42       # AP列
        }

        # DataFrameをA11,B11,F11,J11,W11,AP11に書き込み
        start_row = 11
        for i, row in enumerate(df.itertuples(index=False), start=start_row):
            for col_name, excel_col in column_map.items():
                if col_name in df.columns:
                    value = getattr(row, col_name)
                    ws.cell(row=i, column=excel_col, value=value)

        # バッファに保存（メモリ上）
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.read()
        
        logging.info("テンプレートExcelへの書き込みが完了しました。")

        # --- 5. CSV形式も生成 ---
        # <br>タグを改行に変換
        df_csv = df.copy()
        for col in df_csv.columns:
            df_csv[col] = df_csv[col].astype(str).str.replace('<br>', '\n', regex=False)
        
        csv_buffer = io.StringIO()
        df_csv.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
        
        # --- 6. 全成果物をZIPファイルにまとめる ---
        logging.info("全成果物をZIPファイルにまとめています。")
        base_name = Path(filename).stem
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"{base_name}_構造化設計書.md", md_output_first.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト観点.md", md_output_second.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト仕様書.md", md_output_third.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト仕様書.xlsx", excel_bytes)
            zip_file.writestr(f"{base_name}_テスト仕様書.csv", csv_bytes)
        
        zip_buffer.seek(0)
        zip_bytes = zip_buffer.read()
        logging.info("ZIPファイルの作成が完了しました。")
        
        # --- 7. ユーザーへの返却（ZIPファイル） ---
        # 複数ファイルの場合はファイル数を含む名前を生成
        output_filename = f"テスト仕様書_{len(files)}件.zip"
        encoded_filename = quote(output_filename)
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Type": "application/zip",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return func.HttpResponse(zip_bytes, status_code=200, headers=headers)

    except ValueError as ve:
        logging.error(f"設定エラー: {ve}")
        return func.HttpResponse("サーバー設定エラー", status_code=500)
    except Exception as e:
        logging.error(f"処理全体で予期せぬエラーが発生: {e}")
        return func.HttpResponse("処理中にサーバーエラーが発生しました", status_code=500)


@app.route(route="upload_diff", methods=["POST"])
def upload_diff(req: func.HttpRequest) -> func.HttpResponse:
    try:
        # 新版Excel（必須）
        new_excel_files = req.files.getlist("newExcelFiles")
        if not new_excel_files:
            return func.HttpResponse("新版設計書がアップロードされていません", status_code=400)
        
        # 旧版ファイル（必須）
        old_structured_md_file = req.files.get("oldStructuredMd")
        old_test_spec_md_file = req.files.get("oldTestSpecMd")
        
        if not old_structured_md_file:
            return func.HttpResponse("旧版構造化設計書がアップロードされていません", status_code=400)
        if not old_test_spec_md_file:
            return func.HttpResponse("旧版テスト仕様書がアップロードされていません", status_code=400)
        
        granularity = req.form.get("granularity", "simple")
        if granularity not in ["simple", "detailed"]:
            granularity = "simple"
            
    except Exception as e:
        logging.error(f"ファイル取得エラー: {e}")
        return func.HttpResponse("ファイルの取得に失敗しました", status_code=400)

    logging.info("差分版テスト生成を開始します")

    try:
        # Step 1: 新版Excelを構造化
        logging.info("新版Excelを構造化中...")
        new_structured_md = process_excel_to_markdown(new_excel_files)
        
        # Step 2: 旧版情報の取得（必須）
        logging.info("旧版ファイルを読み込み中...")
        old_structured_md = old_structured_md_file.read().decode('utf-8')
        old_test_spec_md = old_test_spec_md_file.read().decode('utf-8')
        
        # Step 3: 差分検知
        logging.info("差分検知中...")
        diff_prompt = f"""
        【旧版設計書】
        {old_structured_md}
        
        【新版設計書】
        {new_structured_md}
        """
        diff_summary = call_llm(DIFF_DETECTION_PROMPT, diff_prompt)
        logging.info("差分検知完了")
        
        # Step 4: テスト観点抽出（差分考慮）
        logging.info("テスト観点抽出中...")
        test_perspectives_prompt = f"""
        【新版設計書】
        {new_structured_md}
        
        【変更差分】
        {diff_summary}
        """
        test_perspectives = call_llm(EXTRACT_TEST_PERSPECTIVES_PROMPT_WITH_DIFF, test_perspectives_prompt)
        
        # Step 5: テスト仕様書生成（差分・旧版考慮）
        logging.info("テスト仕様書生成中...")
        test_spec_prompt = f"""
        【新版設計書】
        {new_structured_md}
        
        【テスト観点】
        {test_perspectives}
        
        【変更差分】
        {diff_summary}
        
        【旧版テスト仕様書】
        {old_test_spec_md}
        """
        test_spec_md = call_llm(CREATE_TEST_SPEC_PROMPT_WITH_DIFF, test_spec_prompt)
        
        # Step 6: Markdown表をExcelに変換
        logging.info("Excel変換中...")
        md_lines = [line.strip() for line in test_spec_md.splitlines() if line.strip().startswith("|")]
        if not md_lines:
            return func.HttpResponse("テスト仕様書の生成に失敗しました", status_code=500)
        
        tsv_text = "\n".join([line.strip("|").replace("|", "\t") for line in md_lines])
        df = pd.read_csv(io.StringIO(tsv_text), sep="\t")
        df.columns = [col.strip() for col in df.columns]
        
        # 差分版は7列構成（変更種別列あり）または6列構成
        if "変更種別" in df.columns:
            # 変更種別列を除外してExcelに書き込み
            df_for_excel = df.drop(columns=["変更種別"])
        else:
            df_for_excel = df
        
        template_path = "単体テスト仕様書.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        
        column_map = {"No": 1, "大区分": 2, "中区分": 6, "テストケース": 10, "期待結果": 23, "参照元": 42}
        start_row = 11
        for i, row in enumerate(df_for_excel.itertuples(index=False), start=start_row):
            for col_name, excel_col in column_map.items():
                if col_name in df_for_excel.columns:
                    value = getattr(row, col_name)
                    ws.cell(row=i, column=excel_col, value=value)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.read()
        
        # CSV生成
        df_csv = df.copy()
        for col in df_csv.columns:
            df_csv[col] = df_csv[col].astype(str).str.replace('<br>', '\n', regex=False)
        csv_buffer = io.StringIO()
        df_csv.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
        
        # ZIP作成
        logging.info("ZIP作成中...")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("新版_構造化設計書.md", new_structured_md.encode('utf-8'))
            zip_file.writestr("差分サマリー.md", diff_summary.encode('utf-8'))
            zip_file.writestr("テスト観点.md", test_perspectives.encode('utf-8'))
            zip_file.writestr("テスト仕様書.md", test_spec_md.encode('utf-8'))
            zip_file.writestr("テスト仕様書.xlsx", excel_bytes)
            zip_file.writestr("テスト仕様書.csv", csv_bytes)
        
        zip_buffer.seek(0)
        zip_bytes = zip_buffer.read()
        
        output_filename = "テスト仕様書_差分版.zip"
        encoded_filename = quote(output_filename)
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Type": "application/zip",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return func.HttpResponse(zip_bytes, status_code=200, headers=headers)
        
    except Exception as e:
        logging.error(f"差分版処理エラー: {e}")
        return func.HttpResponse("処理中にエラーが発生しました", status_code=500)