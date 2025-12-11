import logging
import io
import pandas as pd
from openpyxl import load_workbook
import re
from pathlib import Path

from core import llm_service

def process_excel_to_markdown(files, progress_callback=None, job_id=None) -> tuple[str, dict]:
    """
    Excelファイル群を構造化されたMarkdownに変換する
    
    Args:
        files: アップロードされたExcelファイルのリスト
        progress_callback: 進捗更新用のコールバック関数
        job_id: ジョブID
    
    Returns:
        tuple[str, dict]: (構造化されたMarkdown, 累積使用量情報)
    """
    all_md_sheets = []  # 各シートの構造化結果を格納
    all_toc_list = []   # 目次用のリンクリスト
    total_usage = {"input_tokens": 0, "output_tokens": 0, "model": ""}
    
    # 総シート数をカウント
    total_sheets = 0
    for file in files:
        file.stream.seek(0)
        file_bytes = file.read()
        excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
        total_sheets += len(excel_data)
    
    logging.info(f"総シート数: {total_sheets}")
    
    # 進捗計算: 10%から40%までを総シート数で均等に分割
    progress_range = 30  # 10%から40%までの範囲
    progress_per_sheet = progress_range / total_sheets if total_sheets > 0 else 0
    current_sheet = 0
    
    for file in files:
        # Excelファイルを読み込む
        file.stream.seek(0)
        file_bytes = file.read()
        filename = file.filename
        excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)
        
        # 各シートを処理
        for sheet_name, df in excel_data.items():
            # シート名をファイル名と組み合わせて一意にする
            full_sheet_name = f"{Path(filename).stem}_{sheet_name}"
            # Markdownアンカー用のIDを生成
            anchor = re.sub(r'[^a-z0-9-]', '', full_sheet_name.strip().lower().replace(' ', '-'))
            all_toc_list.append(f'- [{full_sheet_name}](#{anchor})')
            
            # シートの内容をテキスト化
            sheet_content = f"## {full_sheet_name}\n\n"
            # nanを空文字列に置換し、空白セルを除外
            df_clean = df.fillna('').replace('nan', '')
            raw_text = '\n'.join(df_clean.apply(lambda row: ' | '.join([str(v) for v in row if str(v).strip()]), axis=1))
            structuring_prompt = f'--- Excelシート「{full_sheet_name}」 ---\n{raw_text}'
            
            # LLMで構造化
            try:
                current_sheet += 1
                logging.info(f"「{full_sheet_name}」を構造化... ({current_sheet}/{total_sheets})")
                
                # 進捗更新: 10%から40%までを総シート数で均等に分割
                progress_percent = int(10 + (current_sheet * progress_per_sheet))
                if progress_callback and job_id:
                    progress_callback("structuring", f"設計書を構造化中... ({current_sheet}/{total_sheets}シート)", progress_percent)
                
                structured_content, usage = llm_service.structuring(structuring_prompt)
                total_usage["input_tokens"] += usage["input_tokens"]
                total_usage["output_tokens"] += usage["output_tokens"]
                total_usage["model"] = usage["model"]
                logging.info(f"シート {current_sheet}/{total_sheets} - 入力: {usage['input_tokens']:,}tok, 出力: {usage['output_tokens']:,}tok")
                sheet_content += structured_content
            except Exception as e:
                logging.error(f"AIによるシート「{full_sheet_name}」の構造化中にエラー: {e}")
                sheet_content += "（AIによる構造化に失敗しました）"
                
            all_md_sheets.append(sheet_content)
    
    # 最終的なMarkdownドキュメントを組み立て
    md_output = "# 詳細設計書\n\n## 目次\n\n"
    md_output += "\n".join(all_toc_list)
    md_output += "\n\n---\n\n"
    md_output += "\n\n---\n\n".join(all_md_sheets)
    return md_output, total_usage

def convert_md_to_excel_and_csv(md_output: str, is_diff_mode: bool = False):
    """
    生成されたテスト仕様書（Markdown）をExcelとCSVに変換する
    
    Args:
        md_output: Markdown形式のテスト仕様書
        is_diff_mode: 差分モードかどうか（Trueの場合、変更種別列をExcelから除外）
    
    Returns:
        tuple: (excel_bytes, csv_bytes) ExcelとCSVのバイナリデータ
    """
    # Markdown表の行を抽出（|で始まる行のみ）
    md_lines = [line.strip() for line in md_output.splitlines() if line.strip().startswith("|")]
    
    if not md_lines:
        raise ValueError("テスト仕様書にMarkdown表が見つかりませんでした")
    
    # ヘッダー行の解析（先頭と末尾のパイプを除去）
    header = [h.strip() for h in md_lines[0].strip('|').split('|')]
    
    # 2行目の区切り行（|---|---|）をスキップしてデータ行を抽出
    data_rows = []
    if len(md_lines) > 1 and all(c in '-|: ' for c in md_lines[1]):
        data_rows = md_lines[2:]  # 3行目以降がデータ
    else:
        data_rows = md_lines[1:]  # 区切り行がない場合

    # 各データ行をパース
    data = []
    for row in data_rows:
        data.append([item.strip() for item in row.strip('|').split('|')])

    # pandas DataFrameに変換
    df = pd.DataFrame(data, columns=header)

    # Excel用のDataFrame（差分モードの場合は変更種別列を除外）
    df_for_excel = df.copy()
    if is_diff_mode and "変更種別" in df.columns:
        df_for_excel = df.drop(columns=["変更種別"])

    # --- Excel変換 ---
    # テンプレートファイルを読み込む
    template_path = "単体テスト仕様書.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active
    # 各列のExcel上の位置を定義
    column_map = {
        "No": 1, "大区分": 2, "中区分": 6, "テストケース": 10, "期待結果": 23, "参照元": 42
    }
    start_row = 11  # データの開始行
    # 各データをExcelに書き込む
    for i, row in enumerate(df_for_excel.itertuples(index=False), start=start_row):
        for col_name, excel_col in column_map.items():
            if col_name in df_for_excel.columns:
                value = getattr(row, col_name)
                ws.cell(row=i, column=excel_col, value=value)
    
    # Excelをバイナリデータとして保存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    
    # --- CSV変換 ---
    df_csv = df.copy()
    # <br>タグを改行に変換
    for col in df_csv.columns:
        if df_csv[col].dtype == 'object':
            df_csv[col] = df_csv[col].str.replace('<br>', '\n', regex=False)
    
    # CSVをバイナリデータとして保存（BOM付きUTF-8）
    csv_buffer = io.StringIO()
    df_csv.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
    csv_bytes = csv_buffer.getvalue().encode('utf-8-sig')
    
    return excel_bytes, csv_bytes
